import os
import random
import threading
from datetime import datetime
from flask import Flask
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes
)
from openpyxl import load_workbook

# --- Flask для Render ---
flask_app = Flask(__name__)

@flask_app.route('/')
def home():
    return 'Bot is alive on Render!'

def run_flask():
    flask_app.run(host='0.0.0.0', port=10000)

# --- Глобальные переменные ---
ALL_QUESTIONS = []
MODULES = []
user_progress = {}
user_scores = {}
user_question_sets = {}
user_start_times = {}
leaderboards = {}

# --- Загрузка вопросов ---
def load_all_questions():
    wb = load_workbook("questions.xlsx")
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    return [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

def reload_questions():
    global ALL_QUESTIONS, MODULES
    ALL_QUESTIONS = load_all_questions()
    unique_modules = sorted(set(q["Module"] for q in ALL_QUESTIONS if q.get("Module")))
    MODULES = ["Микс"] + unique_modules

reload_questions()

# --- Команда /start ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_progress[user_id] = 0
    user_scores[user_id] = 0

    buttons = [
        [InlineKeyboardButton(module, callback_data=f"module_{module}")]
        for module in MODULES
    ]
    markup = InlineKeyboardMarkup(buttons)
    await update.message.reply_text(
        "📚 Выберите модуль, по которому хотите пройти викторину:",
        reply_markup=markup
    )

# --- Выбор модуля ---
async def handle_module_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    selected_module = query.data.replace("module_", "")
    context.user_data["selected_module"] = selected_module

    buttons = [
        [InlineKeyboardButton("5 — Разминка 🔄", callback_data="quiz_5")],
        [InlineKeyboardButton("10 — Проверка на прочность 🧠", callback_data="quiz_10")],
        [InlineKeyboardButton("20 — Я ПРО этой игры 🎩", callback_data="quiz_20")]
    ]
    markup = InlineKeyboardMarkup(buttons)
    await query.edit_message_text(
        f"🧠 Вы выбрали модуль: *{selected_module}*\n\nТеперь выберите количество вопросов:",
        parse_mode="Markdown",
        reply_markup=markup
    )

# --- Запуск викторины ---
async def handle_start_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    count = int(query.data.replace("quiz_", ""))
    selected_module = context.user_data.get("selected_module")

    if selected_module == "Микс":
        filtered = ALL_QUESTIONS
    else:
        filtered = [q for q in ALL_QUESTIONS if q.get("Module") == selected_module]

    selected = random.sample(filtered, k=min(count, len(filtered)))

    user_question_sets[user_id] = selected
    user_progress[user_id] = 0
    user_scores[user_id] = 0
    user_start_times[user_id] = datetime.utcnow()
    context.user_data["question_count"] = count

    await query.edit_message_text(
        f"🔥 Загружаю {len(selected)} вопросов из модуля *{selected_module}*.",
        parse_mode="Markdown"
    )
    await send_question(query, context, user_id)

# --- Отправка вопроса ---
async def send_question(source, context, uid):
    chat_id = getattr(source.message, "chat_id", uid)
    idx = user_progress.get(uid, 0)
    questions = user_question_sets.get(uid, [])

    if idx >= len(questions):
        score = user_scores.get(uid, 0)
        total = len(questions)
        duration = (datetime.utcnow() - user_start_times.get(uid, datetime.utcnow())).total_seconds()
        module = context.user_data.get("selected_module", "Микс")
        count = context.user_data.get("question_count", total)

        try:
            user = await context.bot.get_chat(uid)
            name = user.username or user.first_name or f"User{uid}"
        except:
            name = f"User{uid}"

        entry = {
            "uid": uid,
            "name": name,
            "score": score,
            "total": total,
            "duration": duration,
            "date": datetime.utcnow().strftime("%Y-%m-%d")
        }

        key = (module, count)
        board = leaderboards.setdefault(key, [])
        board.append(entry)
        board.sort(key=lambda x: (-x["score"], x["duration"]))
        leaderboards[key] = board[:10]

        await context.bot.send_message(
            chat_id=chat_id,
            text=(
                f"🎉 Викторина завершена!\nВы набрали {score} из {total}.\n"
                f"⏱ Время прохождения: {int(duration)} сек.\n\n"
                "👉 /start — начать заново\n"
                "👉 /score — ваш результат\n"
                "👉 /leaders — таблица лидеров\n"
                "👉 /help — справка"
            )
        )
        return

    q = questions[idx]
    options = {
        "opt_1": q["Option1"],
        "opt_2": q["Option2"],
        "opt_3": q["Option3"],
        "opt_4": q["Option4"]
    }
    context.user_data["correct_option"] = str(q["CorrectAnswer"]).strip()
    context.user_data["options"] = options

    buttons = [
        [InlineKeyboardButton(q["Option1"], callback_data="opt_1")],
        [InlineKeyboardButton(q["Option2"], callback_data="opt_2")],
        [InlineKeyboardButton(q["Option3"], callback_data="opt_3")],
        [InlineKeyboardButton(q["Option4"], callback_data="opt_4")]
    ]
    markup = InlineKeyboardMarkup(buttons)
    await context.bot.send_message(chat_id=chat_id, text=q["Question"], reply_markup=markup)

# --- Обработка ответа ---
async def handle_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    idx = user_progress.get(uid, 0)
    questions = user_question_sets.get(uid, [])

    if idx >= len(questions):
        await query.edit_message_text("Викторина завершена.")
        return

    selected_key = query.data
    selected = context.user_data["options"].get(selected_key)
    correct = context.user_data["correct_option"]
    q = questions[idx]
    explanation = str(q.get("Explanation", "")).strip()

    if selected == correct:
        result = "✅ Верно!"
        user_scores[uid] += 1
    else:
        result = f"❌ Неверно. Правильный ответ: {correct}"

    if explanation:
        result += f"\n\n📘 Объяснение: {explanation}"

    await query.edit_message_text(result)
    user_progress[uid] += 1
    await send_question(query, context, uid)

# --- Команды ---
async def show_score(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    correct = user_scores.get(uid, 0)
    total = user_progress.get(uid, 0)
    await update.message.reply_text(f"📊 Ваш счёт: {correct} из {total} пройденных вопросов.")

async def stop_quiz(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    user_progress.pop(uid, None)
    user_scores.pop(uid, None)
    user_question_sets.pop(uid, None)
    user_start_times.pop(uid, None)
    await update.message.reply_text("🛑 Вы вышли из текущей викторины. Чтобы начать заново, используйте /start.")

HELP_TEXT = (
    "🕵️ Добро пожаловать в викторину по игре *Мафия*!\n\n"
    "📚 Команды:\n"
    "/start — выбрать модуль и начать игру\n"
    "/score — ваш текущий результат\n"
    "/leaders — таблица лидеров\n"
    "/stop — выйти из текущей викторины\n"
    "/help — справка\n\n"
    "Ответы выбираются кнопками. Удачи!"
)

async def show_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(HELP_TEXT, parse_mode="Markdown")

async def show_leaderboard_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("5 — Разминка 🔄", callback_data="leaders_5")],
        [InlineKeyboardButton("10 — Проверка на прочность 🧠", callback_data="leaders_10")],
        [InlineKeyboardButton("20 — Я ПРО этой игры 🎩", callback_data="leaders_20")]
    ]
    markup = InlineKeyboardMarkup(buttons)
    await update.message.reply_text(
        "📊 Выберите количество вопросов, по которым хотите посмотреть таблицу лидеров:",
        reply_markup=markup
    )

async def show_leaderboard_filtered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    count = int(query.data.replace("leaders_", ""))
    lines = []
    found = False

    for (module, total), entries in leaderboards.items():
        if total != count:
            continue
        found = True
        lines.append(f"\n📘 Модуль: *{module}*")
        for rank, entry in enumerate(entries, 1):
            name = entry["name"]
            score = entry["score"]
            duration = int(entry["duration"])
            date = entry["date"]
            lines.append(f"{rank}. @{name}: {score}/{total} — {duration} сек — {date}")

    if not found:
        await query.edit_message_text(f"😕 Пока нет результатов для {count} вопросов.")
    else:
        await query.edit_message_text("🏆 Топ-10 по каждому модулю:\n" + "\n".join(lines), parse_mode="Markdown")

# --- Запуск приложения ---
app = ApplicationBuilder().token(os.getenv("BOT_TOKEN")).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("score", show_score))
app.add_handler(CommandHandler("stop", stop_quiz))
app.add_handler(CommandHandler("help", show_help))
app.add_handler(CommandHandler("leaders", show_leaderboard_prompt))
app.add_handler(CallbackQueryHandler(handle_module_selection, pattern="^module_"))
app.add_handler(CallbackQueryHandler(handle_start_mode, pattern="^quiz_"))
app.add_handler(CallbackQueryHandler(show_leaderboard_filtered, pattern="^leaders_"))
app.add_handler(CallbackQueryHandler(handle_answer))

if __name__ == '__main__':
    threading.Thread(target=run_flask).start()
    app.run_polling()